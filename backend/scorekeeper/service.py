import base64
import json
import re
import sqlite3
from io import BytesIO
from difflib import SequenceMatcher
from pathlib import Path

from groq import Groq
from PIL import Image, ImageEnhance, ImageOps

from .config import get_settings
from .constants import (
    CHAT_MODEL,
    CHAT_SYSTEM_PROMPT_PREFIX,
    DB_FILENAME,
    DEFAULT_ADMIN_CODE,
    FIXED_PLAYERS,
    GROQ_KEY_MISSING_ERROR,
    IMAGE_TOO_LARGE_ERROR,
    MAX_IMAGE_BYTES,
    MATCHABLE_USERNAMES,
    NO_MATCH_DATA_MESSAGE,
    PLAYER_ALIAS_TO_CANONICAL,
    VISION_IMAGE_CONTRAST,
    VISION_IMAGE_MAX_DIMENSION,
    VISION_IMAGE_MIN_DIMENSION,
    VISION_IMAGE_SHARPNESS,
    VISION_IMAGE_UPSCALE_FACTOR,
    VISION_MODEL,
    VISION_CROP_PROMPT,
    VISION_PROMPT,
    VISION_ROW_CROP_BANDS,
)

DB_PATH = Path(__file__).parent / DB_FILENAME


# ── DB helpers ────────────────────────────────────────────────────────────────
def _conn() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con


def _init_db() -> None:
    with _conn() as con:
        con.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key   TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
        """)
        con.execute("""
            CREATE TABLE IF NOT EXISTS players (
                username     TEXT PRIMARY KEY,
                display_name TEXT NOT NULL
            )
        """)
        con.execute("""
            CREATE TABLE IF NOT EXISTS matches (
                id   INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL
            )
        """)
        con.execute("""
            CREATE TABLE IF NOT EXISTS scores (
                username TEXT    NOT NULL,
                match_id INTEGER NOT NULL REFERENCES matches(id),
                points   INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY (username, match_id)
            )
        """)

        # Seed default admin code if not set
        con.execute(
            "INSERT OR IGNORE INTO settings (key, value) VALUES ('admin_code', ?)",
            (DEFAULT_ADMIN_CODE,),
        )

        # Seed fixed players
        for username, display_name in FIXED_PLAYERS:
            con.execute(
                "INSERT OR IGNORE INTO players (username, display_name) VALUES (?,?)",
                (username, display_name),
            )


# ── Fuzzy name → username matching ───────────────────────────────────────────
def _clean(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower())


def _fuzzy_match(raw: str, usernames: list[str]) -> tuple[str, float]:
    """Return (best_username, ratio). ratio >= 0.5 is a match."""
    raw_c = _clean(raw)
    best, best_ratio = usernames[0], 0.0
    for u in usernames:
        ratio = SequenceMatcher(None, raw_c, _clean(u)).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best = u
    return best, best_ratio


def _canonical_username(username: str) -> str:
    return PLAYER_ALIAS_TO_CANONICAL.get(username, username)


def _prepare_image_for_vision(image_bytes: bytes) -> bytes:
    """Lightly upscale and sharpen screenshots so small leaderboard text is easier to read."""
    with Image.open(BytesIO(image_bytes)) as img:
        img = ImageOps.exif_transpose(img).convert("RGB")
        short_dim = min(img.size)
        max_dim = max(img.size)
        if short_dim < VISION_IMAGE_MIN_DIMENSION:
            scale = min(
                VISION_IMAGE_UPSCALE_FACTOR,
                VISION_IMAGE_MAX_DIMENSION / max_dim,
                VISION_IMAGE_MIN_DIMENSION / short_dim,
            )
            if scale > 1:
                new_size = (
                    max(1, int(round(img.width * scale))),
                    max(1, int(round(img.height * scale))),
                )
                img = img.resize(new_size, Image.Resampling.LANCZOS)
        elif max_dim > VISION_IMAGE_MAX_DIMENSION:
            img.thumbnail(
                (VISION_IMAGE_MAX_DIMENSION, VISION_IMAGE_MAX_DIMENSION),
                Image.Resampling.LANCZOS,
            )

        img = ImageEnhance.Contrast(img).enhance(VISION_IMAGE_CONTRAST)
        img = ImageEnhance.Sharpness(img).enhance(VISION_IMAGE_SHARPNESS)

        buf = BytesIO()
        img.save(buf, format="PNG", optimize=True)
        return buf.getvalue()


def _prepare_image_crop_for_vision(image_bytes: bytes, crop_box: tuple[float, float, float, float] | None = None) -> bytes:
    with Image.open(BytesIO(image_bytes)) as img:
        img = ImageOps.exif_transpose(img).convert("RGB")
        if crop_box is not None:
            left = int(round(img.width * crop_box[0]))
            top = int(round(img.height * crop_box[1]))
            right = int(round(img.width * crop_box[2]))
            bottom = int(round(img.height * crop_box[3]))
            img = img.crop((left, top, right, bottom))

        short_dim = min(img.size)
        max_dim = max(img.size)
        if short_dim < VISION_IMAGE_MIN_DIMENSION:
            scale = min(
                VISION_IMAGE_UPSCALE_FACTOR,
                VISION_IMAGE_MAX_DIMENSION / max_dim,
                VISION_IMAGE_MIN_DIMENSION / short_dim,
            )
            if scale > 1:
                new_size = (
                    max(1, int(round(img.width * scale))),
                    max(1, int(round(img.height * scale))),
                )
                img = img.resize(new_size, Image.Resampling.LANCZOS)
        elif max_dim > VISION_IMAGE_MAX_DIMENSION:
            img.thumbnail(
                (VISION_IMAGE_MAX_DIMENSION, VISION_IMAGE_MAX_DIMENSION),
                Image.Resampling.LANCZOS,
            )

        img = ImageEnhance.Contrast(img).enhance(VISION_IMAGE_CONTRAST)
        img = ImageEnhance.Sharpness(img).enhance(VISION_IMAGE_SHARPNESS)

        buf = BytesIO()
        img.save(buf, format="PNG", optimize=True)
        return buf.getvalue()


def _vision_image_variants(image_bytes: bytes) -> list[bytes]:
    variants = [_prepare_image_for_vision(image_bytes)]
    # Overlapping row crops give the model a much larger effective zoom on each leaderboard line.
    variants.extend(_prepare_image_crop_for_vision(image_bytes, band) for band in VISION_ROW_CROP_BANDS)
    return variants


def _extract_players_from_result(
    result: dict,
    matchable_usernames: list[str],
) -> dict[str, tuple[float, str, float]]:
    extracted: dict[str, tuple[float, str, float]] = {}
    for p in result.get("players", []):
        name = str(p.get("name", "")).strip()
        pts_raw = str(p.get("points", "0")).strip().replace(",", "")
        if not name or not re.fullmatch(r"-?\d+(?:\.\d+)?", pts_raw):
            continue
        username, ratio = _fuzzy_match(name, matchable_usernames)
        username = _canonical_username(username)
        if ratio < 0.5:
            continue

        try:
            points = float(pts_raw)
        except ValueError:
            continue
        current = extracted.get(username)
        if current is None or ratio > current[2] or (ratio == current[2] and current[0] == 0 and points != 0):
            extracted[username] = (points, name, ratio)
    return extracted


def _build_vision_prompt(base_prompt: str, candidate_usernames: list[str]) -> str:
    candidates = "\n".join(f"- {username}" for username in candidate_usernames)
    return (
        f"{base_prompt}\n"
        "Choose names from this exact candidate list whenever possible:\n"
        f"{candidates}\n"
        "Prefer the closest visible candidate if the text is blurry.\n"
        "Do not invent usernames outside the list."
    )


# ── Service ───────────────────────────────────────────────────────────────────
class ScoreKeeperService:
    def __init__(self):
        _init_db()
        self._usernames = [u for u, _ in FIXED_PLAYERS]
        self._matchable_usernames = MATCHABLE_USERNAMES
        self._groq: Groq | None = None

    def _get_groq(self) -> Groq:
        api_key = get_settings().groq_api_key.strip()
        if not api_key:
            raise RuntimeError(GROQ_KEY_MISSING_ERROR)
        if self._groq is None:
            self._groq = Groq(api_key=api_key)
        return self._groq

    # ── Admin code ─────────────────────────────────────────────────────────

    def verify_code(self, code: str) -> bool:
        with _conn() as con:
            row = con.execute("SELECT value FROM settings WHERE key='admin_code'").fetchone()
        return row and row["value"] == code

    def change_code(self, old_code: str, new_code: str) -> tuple[bool, str]:
        if not self.verify_code(old_code):
            return False, "Galat purana code hai. / Old code is incorrect."
        if len(new_code) < 4:
            return False, "New code must be at least 4 characters."
        with _conn() as con:
            con.execute("UPDATE settings SET value=? WHERE key='admin_code'", (new_code,))
        return True, "Code successfully changed!"

    # ── Image extraction ────────────────────────────────────────────────────

    def extract_from_image(self, image_bytes: bytes, content_type: str) -> list[dict]:
        """
        Call vision LLM, fuzzy-map names to the 8 fixed players.
        Returns list of {username, display_name, raw_name, points} for ALL 8 players.
        """
        if len(image_bytes) > MAX_IMAGE_BYTES:
            raise ValueError(IMAGE_TOO_LARGE_ERROR)
        groq = self._get_groq()
        extracted: dict[str, tuple[float, str, float]] = {}
        for variant_index, prepared_bytes in enumerate(_vision_image_variants(image_bytes)):
            b64 = base64.b64encode(prepared_bytes).decode()
            prompt = _build_vision_prompt(
                VISION_PROMPT if variant_index == 0 else VISION_CROP_PROMPT,
                self._matchable_usernames,
            )
            completion = groq.chat.completions.create(
                model=VISION_MODEL,
                temperature=0,
                messages=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:image/png;base64,{b64}"},
                        },
                        {
                            "type": "text",
                            "text": prompt,
                        },
                    ],
                }],
                max_tokens=1024,
            )

            raw = re.sub(r"```(?:json)?", "", completion.choices[0].message.content or "").strip()
            try:
                result = json.loads(raw)
            except Exception:
                m = re.search(r"\{.*\}", raw, re.DOTALL)
                result = json.loads(m.group()) if m else {}

            variant_extracted = _extract_players_from_result(result, self._matchable_usernames)
            for username, candidate in variant_extracted.items():
                current = extracted.get(username)
                if current is None or candidate[2] > current[2] or (
                    candidate[2] == current[2] and current[0] == 0 and candidate[0] != 0
                ):
                    extracted[username] = candidate

        # Build full list of all 8 players
        player_map = {u: dn for u, dn in FIXED_PLAYERS}
        return [
            {
                "username":     u,
                "display_name": player_map[u],
                "raw_name":     extracted.get(u, (0, "—", 0.0))[1],
                "points":       extracted.get(u, (0, "—", 0.0))[0],
            }
            for u in self._usernames
        ]

    # ── Save confirmed match ────────────────────────────────────────────────

    def save_match(self, player_scores: list[dict]) -> tuple[str, int]:
        """player_scores: [{username, points}]"""
        with _conn() as con:
            match_number = con.execute("SELECT COALESCE(MAX(id), 0) + 1 FROM matches").fetchone()[0]
            match_name = f"Match {match_number}"
            cur = con.execute("INSERT INTO matches (name) VALUES (?)", (match_name,))
            match_id = cur.lastrowid
            for ps in player_scores:
                username = _canonical_username(ps["username"])
                con.execute(
                    "INSERT INTO scores (username, match_id, points) VALUES (?,?,?)"
                    " ON CONFLICT(username, match_id) DO UPDATE SET points=excluded.points",
                    (username, match_id, ps["points"]),
                )
        return match_name, match_number

    def list_matches(self) -> list[dict]:
        with _conn() as con:
            rows = con.execute("SELECT id, name FROM matches ORDER BY id").fetchall()
        return [dict(r) for r in rows]

    def get_match(self, match_id: int) -> dict:
        with _conn() as con:
            match = con.execute("SELECT id, name FROM matches WHERE id=?", (match_id,)).fetchone()
            if match is None:
                raise ValueError("Match not found.")
            scores = con.execute(
                "SELECT username, points FROM scores WHERE match_id=?",
                (match_id,),
            ).fetchall()

        score_map: dict[str, int] = {}
        for row in scores:
            score_map[_canonical_username(row["username"])] = row["points"]
        player_map = {u: dn for u, dn in FIXED_PLAYERS}
        players = [
            {
                "username": u,
                "display_name": player_map[u],
                "points": score_map.get(u, 0),
            }
            for u in self._usernames
        ]
        return {
            "match_id": match["id"],
            "match_name": match["name"],
            "players": players,
        }

    def update_match(self, match_id: int, player_scores: list[dict]) -> tuple[str, int]:
        with _conn() as con:
            match = con.execute("SELECT name FROM matches WHERE id=?", (match_id,)).fetchone()
            if match is None:
                raise ValueError("Match not found.")
            for ps in player_scores:
                username = _canonical_username(ps["username"])
                con.execute(
                    "INSERT INTO scores (username, match_id, points) VALUES (?,?,?) "
                    "ON CONFLICT(username, match_id) DO UPDATE SET points=excluded.points",
                    (username, match_id, ps["points"]),
                )
        return match["name"], match_id

    def delete_match(self, match_id: int) -> str:
        with _conn() as con:
            match = con.execute("SELECT name FROM matches WHERE id=?", (match_id,)).fetchone()
            if match is None:
                raise ValueError("Match not found.")
            con.execute("DELETE FROM scores WHERE match_id=?", (match_id,))
            con.execute("DELETE FROM matches WHERE id=?", (match_id,))
        return match["name"]

    def reset_data(self) -> None:
        with _conn() as con:
            con.execute("DELETE FROM scores")
            con.execute("DELETE FROM matches")

    # ── Standings ───────────────────────────────────────────────────────────

    def get_standings(self) -> dict:
        with _conn() as con:
            matches = [dict(r) for r in con.execute("SELECT id, name FROM matches ORDER BY id")]
            scores = con.execute("SELECT username, match_id, points FROM scores").fetchall()
            players = [dict(r) for r in con.execute("SELECT username, display_name FROM players")]

        score_map: dict[str, dict[int, int]] = {}
        for row in scores:
            username = _canonical_username(row["username"])
            score_map.setdefault(username, {})[row["match_id"]] = row["points"]

        standings = []
        for p in players:
            u = p["username"]
            sm = score_map.get(u, {})
            total = sum(sm.values())
            match_points = {m["name"]: sm.get(m["id"], 0) for m in matches}
            standings.append({
                "username":     u,
                "display_name": p["display_name"],
                "total":        total,
                "matches":      match_points,
            })

        standings.sort(key=lambda x: x["total"], reverse=True)
        for i, s in enumerate(standings):
            s["rank"] = i + 1

        return {"players": standings, "match_headers": [m["name"] for m in matches]}

    # ── Export ──────────────────────────────────────────────────────────────

    def export_xlsx(self) -> bytes:
        import io
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        data = self.get_standings()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Scores"

        headers = ["Rank", "Username", "Display Name"] + data["match_headers"] + ["Total"]
        hfill = PatternFill("solid", fgColor="1e293b")
        hfont = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hfill
            c.font = hfont
            c.alignment = Alignment(horizontal="center")

        for ri, p in enumerate(data["players"], 2):
            ws.cell(row=ri, column=1, value=p["rank"])
            ws.cell(row=ri, column=2, value=p["username"])
            ws.cell(row=ri, column=3, value=p["display_name"])
            for ci, mh in enumerate(data["match_headers"]):
                ws.cell(row=ri, column=4 + ci, value=p["matches"].get(mh, 0))
            ws.cell(row=ri, column=len(headers), value=p["total"])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Chat ────────────────────────────────────────────────────────────────

    def chat(self, question: str) -> str:
        groq = self._get_groq()
        data = self.get_standings()
        if not data["match_headers"]:
            return NO_MATCH_DATA_MESSAGE

        lines = ["=== Match Points Data ==="]
        for p in data["players"]:
            details = " | ".join(f"{m}: {v}" for m, v in p["matches"].items())
            lines.append(
                f"Rank {p['rank']}: {p['display_name']} (@{p['username']}) "
                f"— Total: {p['total']} | {details}"
            )

        completion = groq.chat.completions.create(
            model=CHAT_MODEL,
            messages=[
                {
                    "role": "system",
                    "content": CHAT_SYSTEM_PROMPT_PREFIX + "\n".join(lines),
                },
                {"role": "user", "content": question},
            ],
            max_tokens=512,
        )
        return completion.choices[0].message.content or "Sorry, could not generate a response."


_service: ScoreKeeperService | None = None


def get_service() -> ScoreKeeperService:
    global _service
    if _service is None:
        _service = ScoreKeeperService()
    return _service
